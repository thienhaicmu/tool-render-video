from pathlib import Path
from openpyxl import Workbook, load_workbook


def append_rows(xlsx_path: Path, headers: list[str], rows: list[list]):
    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(xlsx_path)
